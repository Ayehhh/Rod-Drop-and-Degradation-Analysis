import os
import io
import shutil
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
from scipy import stats
from scipy.optimize import brentq, curve_fit
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from datetime import datetime, timedelta
import streamlit as st

# Set Streamlit Page Config
st.set_page_config(
    page_title="Rod Drop Prognostic Analysis",
    page_icon="⚙️",
    layout="wide"
)

st.title("⚙️ Reciprocating Compressor Rod Drop Prognostic Analysis")
st.markdown("Model wear degradation and project alarm breach dates for piston rod drop.")

# ==========================================
# 1. DATA SOURCE & PARAMETERS CONFIGURATION
# ==========================================
st.sidebar.header("1. Data Source")
data_source = st.sidebar.selectbox(
    "Choose Analysis Mode:",
    ["User Specified", "Sample Data"]
)

df = None

if data_source == "User Specified":
    st.sidebar.subheader("Upload Dataset")
    uploaded_file = st.sidebar.file_uploader("Upload Excel Dataset (.xlsx / .xls)", type=["xlsx", "xls"])
    
    st.sidebar.subheader("Engineering Parameters")
    NEW_CLEARANCE = st.sidebar.number_input("As-New Clearance [mm]", value=2.000, step=0.001, format="%.3f")
    BN_L_THRESHOLD = st.sidebar.number_input("Bently Nevada L Alarm Threshold [µm]", value=-1500.0, step=10.0, format="%.1f")
    BN_LL_THRESHOLD = st.sidebar.number_input("Bently Nevada LL Alarm Threshold [µm]", value=-2000.0, step=10.0, format="%.1f")
    MIN_CLEARANCE = st.sidebar.number_input("Minimum Allowable Clearance / OEM Limit [mm]", value=0.500, step=0.001, format="%.3f")
    CONFIDENCE_PCT = st.sidebar.number_input("Confidence Level Analysis [%]", value=95.0, min_value=50.0, max_value=99.9, step=1.0)

    if uploaded_file is not None:
        df = pd.read_excel(uploaded_file)
    else:
        st.info("👈 Please upload an Excel dataset in the sidebar to begin analysis.")
        st.stop()

else:  # Sample Data Mode
    st.sidebar.success("✅ Running in Sample Data Mode")
    
    # Preset sample parameters
    NEW_CLEARANCE = 2.000
    BN_L_THRESHOLD = -1500.0
    BN_LL_THRESHOLD = -2000.0
    MIN_CLEARANCE = 0.500
    
    st.sidebar.subheader("Sample Threshold Values")
    st.sidebar.text(f"• As-New Clearance: {NEW_CLEARANCE:.3f} mm")
    st.sidebar.text(f"• L Alarm: {BN_L_THRESHOLD:.1f} µm")
    st.sidebar.text(f"• LL Alarm: {BN_LL_THRESHOLD:.1f} µm")
    st.sidebar.text(f"• OEM Min Clearance: {MIN_CLEARANCE:.3f} mm")
    
    CONFIDENCE_PCT = st.sidebar.number_input("Confidence Level Analysis [%]", value=95.0, min_value=50.0, max_value=99.9, step=1.0)

    # Generate realistic historical sample data automatically
    np.random.seed(42)
    dates = pd.date_range(end=datetime.now(), periods=18, freq='30D')
    # Wear progression from 0 to -1100 um with slight measurement noise
    days_passed = np.arange(18) * 30
    synthetic_wear = 0.05 * (days_passed ** 1.3) + np.random.normal(0, 15, 18)
    raw_readings = -np.abs(synthetic_wear)
    
    df = pd.DataFrame({
        "timestamp": dates,
        "raw_um": raw_readings
    })

# Setup Output Directory
OUTPUT_DIR = "Prognosis_Output_Files"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Calculate Target Clearances and Wear Limits
CLEARANCE_AT_L = NEW_CLEARANCE - (abs(BN_L_THRESHOLD) / 1000.0)
CLEARANCE_AT_LL = NEW_CLEARANCE - (abs(BN_LL_THRESHOLD) / 1000.0)
WEAR_TARGET_L = abs(BN_L_THRESHOLD)
WEAR_TARGET_LL = abs(BN_LL_THRESHOLD)
WEAR_TARGET_MIN = (NEW_CLEARANCE - MIN_CLEARANCE) * 1000.0

# ==========================================
# 2. DATA PREPROCESSING
# ==========================================
if len(df.columns) >= 2:
    df = df.iloc[:, :2]
    df.columns = ["timestamp", "raw_um"]

df["timestamp"] = pd.to_datetime(df["timestamp"])
df["raw_um"] = pd.to_numeric(df["raw_um"], errors="coerce")
df = df.dropna().sort_values("timestamp").reset_index(drop=True)

df["wear_um"] = -df["raw_um"]
df["clearance_mm"] = NEW_CLEARANCE - (df["wear_um"] / 1000.0)

t0 = df["timestamp"].iloc[0]
days = (df["timestamp"] - t0).dt.total_seconds().values / 86400.0
wear_um = df["wear_um"].values

latest_wear_um = df["wear_um"].iloc[-1]
latest_clearance_mm = df["clearance_mm"].iloc[-1]

# ==========================================
# 3. REGRESSION MODELING
# ==========================================
def _lin(x, a, b): return a * x + b
def _quad(x, a, b, c): return a * x**2 + b * x + c
def _power(x, a, b): return a * np.power(np.maximum(x, 1e-6), b)
def _expo(x, a, b): return a * np.exp(np.clip(b * x, -100, 100))
def _logf(x, a, b): return a * np.log(x + 1.0) + b

MODELS = {
    "Linear": (_lin, [1.0, 0.0]),
    "Quadratic": (_quad, [0.01, 1.0, 0.0]),
    "Power": (_power, [1.0, 1.0]),
    "Exponential": (_expo, [1.0, 0.01]),
    "Logarithmic": (_logf, [1.0, 0.0])
}

model_results = {}
for name, (func, p0) in MODELS.items():
    try:
        popt, _ = curve_fit(func, days, wear_um, p0=p0, maxfev=10000)
        y_pred = func(days, *popt)
        ss_res = np.sum((wear_um - y_pred)**2)
        ss_tot = np.sum((wear_um - np.mean(wear_um))**2)
        r2 = 1 - (ss_res / ss_tot) if ss_tot > 0 else 0
        dof = max(len(days) - len(popt), 1)
        resid_std = np.sqrt(ss_res / dof)
        model_results[name] = {
            "func": func, "popt": popt, "r2": r2,
            "resid_std": resid_std, "dof": dof, "ss_res": ss_res
        }
    except Exception:
        pass

best_name = max(model_results, key=lambda k: model_results[k]["r2"])
best = model_results[best_name]

# ==========================================
# 4. PROGNOSTIC BREACH CALCULATION
# ==========================================
def solve_crossing(model, target_wear, conf_pct, max_days=3650):
    func, popt, dof, std = model["func"], model["popt"], model["dof"], model["resid_std"]
    t_val = stats.t.ppf((1 + conf_pct / 100.0) / 2, dof) if dof >= 1 else 0.0
    band = t_val * std

    def get_date(f):
        xs = np.linspace(0, max_days, 15000)
        ys = f(xs)
        idx = np.where(np.diff(np.sign(ys)) != 0)[0]
        if len(idx) == 0: return None
        try: return brentq(f, xs[idx[0]], xs[idx[0]+1])
        except: return None

    early = get_date(lambda x: func(x, *popt) + band - target_wear)
    central = get_date(lambda x: func(x, *popt) - target_wear)
    late = get_date(lambda x: func(x, *popt) - band - target_wear)
    return early, central, late

max_horizon = max(days[-1] * 4, days[-1] + 3650)
f_L = solve_crossing(best, WEAR_TARGET_L, CONFIDENCE_PCT, max_horizon)
f_LL = solve_crossing(best, WEAR_TARGET_LL, CONFIDENCE_PCT, max_horizon)
f_Min = solve_crossing(best, WEAR_TARGET_MIN, CONFIDENCE_PCT, max_horizon)

# Display Key Metrics
m1, m2, m3, m4 = st.columns(4)
m1.metric("Selected Model", f"{best_name}", f"R² = {best['r2']:.4f}")
m2.metric("Latest Wear", f"{latest_wear_um:.1f} µm")
m3.metric("Current Clearance", f"{latest_clearance_mm:.3f} mm")
m4.metric("Confidence Level", f"{CONFIDENCE_PCT:.1f}%")

st.subheader("📋 Prognostic Breach Projection Summary")
prognosis_data = []
for label, (e, c, l), target_c in [("L Alarm", f_L, CLEARANCE_AT_L), ("LL Alarm", f_LL, CLEARANCE_AT_LL), ("OEM Min Limit", f_Min, MIN_CLEARANCE)]:
    c_date = (t0 + timedelta(days=c)).strftime('%Y-%m-%d') if c else "N/A"
    e_date = (t0 + timedelta(days=e)).strftime('%Y-%m-%d') if e else "N/A"
    l_date = (t0 + timedelta(days=l)).strftime('%Y-%m-%d') if l else "N/A"
    prognosis_data.append({
        "Threshold Level": label,
        "Target Clearance (mm)": f"{target_c:.3f}",
        "Earliest Predicted Date": e_date,
        "Expected Predicted Date": c_date,
        "Latest Predicted Date": l_date
    })
st.table(pd.DataFrame(prognosis_data))

# ==========================================
# 5. GENERATE VISUALIZATIONS
# ==========================================
st.subheader("📈 Prognostic Trend & Cross-Section Schematic")
col1, col2 = st.columns(2)

# Trend Chart
fig, ax = plt.subplots(figsize=(8, 5), dpi=150)
ax.scatter(df["timestamp"], df["clearance_mm"], color="#1f77b4", s=22, alpha=0.8, label="Measured Observations")

candidate_days = [d for d in [f_L[1], f_LL[1], f_Min[1]] if d is not None]
x_max_plot = max(candidate_days) * 1.15 if candidate_days else days[-1] * 2
x_plot = np.linspace(0, x_max_plot, 500)
y_wear_plot = best["func"](x_plot, *best["popt"])
y_clear_plot = NEW_CLEARANCE - (y_wear_plot / 1000.0)
dates_plot = [t0 + timedelta(days=d) for d in x_plot]

t_val = stats.t.ppf((1 + CONFIDENCE_PCT / 100.0) / 2, best["dof"]) if best["dof"] >= 1 else 0.0
band_mm = (t_val * best["resid_std"]) / 1000.0

ax.plot(dates_plot, y_clear_plot, color="#d62728", linewidth=2, label=f"Best Fit ({best_name})")
ax.fill_between(dates_plot, y_clear_plot - band_mm, y_clear_plot + band_mm, color="#d62728", alpha=0.15, label=f"{CONFIDENCE_PCT:.0f}% CI")

ax.axhline(CLEARANCE_AT_L, color="#ff7f0e", linestyle="--", linewidth=1.5, label=f"L Alarm ({CLEARANCE_AT_L:.3f} mm)")
ax.axhline(CLEARANCE_AT_LL, color="#d62728", linestyle="--", linewidth=1.5, label=f"LL Alarm ({CLEARANCE_AT_LL:.3f} mm)")
ax.axhline(MIN_CLEARANCE, color="black", linestyle=":", linewidth=1.5, label=f"Min Clearance ({MIN_CLEARANCE:.3f} mm)")

ax.set_ylabel("Clearance (mm)")
ax.set_xlabel("Date")
ax.set_title("Piston Rod Clearance Prognostic Trend", fontweight="bold")
ax.legend(loc="lower left", fontsize=7)
ax.grid(True, linestyle=":", alpha=0.6)
fig.autofmt_xdate()
fig.tight_layout()

plot_img_path = os.path.join(OUTPUT_DIR, "clearance_trend_plot_en.png")
fig.savefig(plot_img_path, dpi=150, bbox_inches="tight")
col1.pyplot(fig)

# Schematic Drawing
def generate_piston_head_schematic(new_clear, current_clear, l_clear, ll_clear, min_clear, current_wear):
    fig_s, ax_s = plt.subplots(figsize=(8, 6), dpi=150)
    ax_s.set_xlim(-7.5, 7.5)
    ax_s.set_ylim(-6.5, 4.5)
    ax_s.axis("off")

    liner_radius = 3.2
    liner_center_y = 0.5
    liner = patches.Circle((0, liner_center_y), liner_radius, linewidth=3, edgecolor="#1B365D", facecolor="#F4F6F9", zorder=1)
    ax_s.add_patch(liner)

    ax_s.axhline(liner_center_y, color="#BDC3C7", linestyle=":", linewidth=1, zorder=2)
    ax_s.axvline(0, color="#BDC3C7", linestyle=":", linewidth=1, zorder=2)

    piston_radius = 2.6
    y_piston_center = liner_center_y - (current_wear / 1000.0) * 0.8
    piston = patches.Circle((0, y_piston_center), piston_radius, linewidth=2.5, edgecolor="#2C3E50", facecolor="#BDC3C7", zorder=3)
    ax_s.add_patch(piston)
    ax_s.text(0, y_piston_center + 0.6, "PISTON HEAD\n(Front View)", fontsize=9, fontweight="bold", color="#1B365D", ha="center", va="center", zorder=4)

    rider_ring = patches.Wedge((0, y_piston_center), piston_radius, 225, 315, width=0.35, edgecolor="#D35400", facecolor="#E67E22", zorder=4)
    ax_s.add_patch(rider_ring)

    y_as_new = -2.8
    y_current = -3.4
    y_L = -4.0
    y_LL = -4.6
    y_min = -5.2

    ax_s.hlines(y=y_as_new, xmin=-5.5, xmax=-0.3, colors="#27AE60", linestyles="--", linewidth=1.5, zorder=5)
    ax_s.hlines(y=y_current, xmin=-5.5, xmax=0.3, colors="#2980B9", linestyles="-.", linewidth=1.8, zorder=5)
    ax_s.hlines(y=y_L, xmin=0.3, xmax=5.5, colors="#F39C12", linestyles="--", linewidth=1.5, zorder=5)
    ax_s.hlines(y=y_LL, xmin=0.3, xmax=5.5, colors="#E74C3C", linestyles="--", linewidth=1.5, zorder=5)
    ax_s.hlines(y=y_min, xmin=-5.5, xmax=5.5, colors="#8E44AD", linestyles="-", linewidth=2.0, zorder=5)

    ax_s.text(-5.7, y_as_new, f"As-New: {new_clear:.3f} mm", fontsize=8, color="#27AE60", ha="right", va="center")
    ax_s.text(-5.7, y_current, f"Est. Current: {current_clear:.3f} mm", fontsize=8, color="#2980B9", ha="right", va="center")
    ax_s.text(5.7, y_L, f"L Alarm: {l_clear:.3f} mm", fontsize=8, color="#F39C12", ha="left", va="center")
    ax_s.text(5.7, y_LL, f"LL Alarm: {ll_clear:.3f} mm", fontsize=8, color="#E74C3C", ha="left", va="center")
    ax_s.text(5.7, y_min, f"OEM Min: {min_clear:.3f} mm", fontsize=8, color="#8E44AD", ha="left", va="center")

    ax_s.set_title("Cross-Section Clearance Schematic", fontsize=11, fontweight="bold")
    plt.tight_layout()
    schematic_img_path = os.path.join(OUTPUT_DIR, "piston_head_schematic.png")
    fig_s.savefig(schematic_img_path, dpi=150, bbox_inches="tight")
    plt.close()
    return fig_s, schematic_img_path

fig_schematic, schematic_img_path = generate_piston_head_schematic(NEW_CLEARANCE, latest_clearance_mm, CLEARANCE_AT_L, CLEARANCE_AT_LL, MIN_CLEARANCE, latest_wear_um)
col2.pyplot(fig_schematic)

# ==========================================
# 6. EXCEL REPORT GENERATION & DOWNLOAD
# ==========================================
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Summary & Prognosis"
ws1.views.sheetView[0].showGridLines = True

NAVY_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
STEEL_FILL = PatternFill(start_color="4A777A", end_color="4A777A", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
REGULAR_FONT = Font(name="Calibri", size=11)

ws1.merge_cells("B2:G2")
ws1["B2"] = "DEGRADATION AND PROGNOSTIC ANALYSIS REPORT"
ws1["B2"].font = TITLE_FONT
ws1["B2"].fill = NAVY_FILL
ws1["B2"].alignment = Alignment(horizontal="center", vertical="center")

params = [
    ("As-Left Bottom Clearance", NEW_CLEARANCE, "mm"),
    ("Bently Nevada L Alarm Threshold", BN_L_THRESHOLD, "um"),
    ("Bently Nevada LL Alarm Threshold", BN_LL_THRESHOLD, "um"),
    ("Minimum Allowable Clearance (OEM)", MIN_CLEARANCE, "mm"),
    ("Calculated L Clearance Limit", CLEARANCE_AT_L, "mm"),
    ("Calculated LL Clearance Limit", CLEARANCE_AT_LL, "mm"),
    ("Statistical Confidence Level", CONFIDENCE_PCT / 100.0, "%")
]

for idx, (param, val, unit) in enumerate(params, start=5):
    ws1.cell(row=idx, column=2, value=param).font = BOLD_FONT
    cell = ws1.cell(row=idx, column=3, value=val)
    cell.font = REGULAR_FONT
    cell.alignment = Alignment(horizontal="right")
    if unit == "mm": cell.number_format = "0.000"
    elif unit == "um": cell.number_format = "0.0"
    elif unit == "%": cell.number_format = "0.0%"
    ws1.cell(row=idx, column=4, value=unit).font = REGULAR_FONT

img_trend = Image(plot_img_path)
ws1.add_image(img_trend, "B14")

img_schematic = Image(schematic_img_path)
ws1.add_image(img_schematic, "I14")

excel_buffer = io.BytesIO()
wb.save(excel_buffer)
excel_buffer.seek(0)

zip_path = f"{OUTPUT_DIR}.zip"
shutil.make_archive(OUTPUT_DIR, 'zip', OUTPUT_DIR)
with open(zip_path, "rb") as f:
    zip_bytes = f.read()

st.subheader("📥 Download Prognosis Reports")
dcol1, dcol2 = st.columns(2)
dcol1.download_button(
    label="📄 Download Excel Summary Report",
    data=excel_buffer,
    file_name=f"Piston_Rod_Clearance_Prognostic_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
dcol2.download_button(
    label="📦 Download Complete ZIP Package (Report + Images)",
    data=zip_bytes,
    file_name=f"Prognosis_Output_Files_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
    mime="application/zip"
)
